import logging
import hashlib
import json
import openpyxl
import os
from pathlib import Path

BUF_SIZE = 65536


class ExcelFile:
    name        = ""
    __filename  = ""
    __hash_file = ""
    __wb_obj = None

    def __init__(self, filename):
        self.set_filename(filename)
        name = os.path.basename(filename)
        self.set_name(name)
        self.__setup_hash()

    def set_name(self, name):
        self.name = name

    def set_filename(self, filename):
        self.__filename = filename

    def get_hash(self):
        return self.__hash_file

    def __setup_hash(self):
        hash_file = self.__calculate_hash()
        self.__hash_file = hash_file

    def __calculate_hash(self):
        sha1 = hashlib.sha1()
        with open(self.__filename, 'rb') as f:
            while True:
                data = f.read(BUF_SIZE)
                if not data:
                    break
                sha1.update(data)
        return sha1.hexdigest()

    def to_dict(self):
        return dict(name=self.name, hash=self.__hash_file, content=self.get_content())

    def __str__(self):
        return json.dumps(self.to_dict())

    def __load_file(self):
        xlsx_file = Path('.', self.__filename)
        self.__wb_obj = openpyxl.load_workbook(xlsx_file)

    def get_content(self):
        self.__load_file()
        
        file_contents = []
        for ws in self.__wb_obj.worksheets:
            rows_contents = []
            for row in ws.iter_rows():
                content = self.__get_row_content(row)
                if len(content) > 0:
                    rows_contents.append(content)
            if len(rows_contents) > 0:
                file_contents.append(dict(title=ws.title, content=rows_contents))
                
            logging.info("max_columns={max_columns}, max_row={max_rows}".format(max_columns=ws.max_column, max_rows=ws.max_column))
        return file_contents

    def __get_row_content(self, row):
        content = []
        key = None
        for col in row:
            if col.value:
                if key:
                    obj = {}
                    obj[key.rstrip(":")] = col.value
                    content.append(obj)
                    key = None
                    continue
                if ":" in col.value:
                    key = col.value
                    continue
                content.append(col.value)
        return content

    def print_content(self):
        self.__load_file()
        print(self.get_content())